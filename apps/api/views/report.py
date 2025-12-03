from django.http import HttpResponse
from django.utils.dateparse import parse_date
from rest_framework.views import APIView
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.utils import get_column_letter
from django.utils.encoding import iri_to_uri
from apps.api.core.permissions import permissions
from apps.application.models import Application, ApplicationStatus, Form

@permissions('r: user; 3p: admin, proxy')
class ReportXlsxView(APIView):
    def get(self, request):
        dept_obj = getattr(request.user, 'department', None)
        dept_id = getattr(dept_obj, 'id', None)
        dept_name = getattr(dept_obj, 'name', 'Unknown department')
        date_from = request.GET.get('dateFrom')
        date_to = request.GET.get('dateTo')

        if dept_id is None or not date_from or not date_to:
            return HttpResponse('Missing query parameters or user department', status=400)

        d_from = parse_date(date_from)
        d_to = parse_date(date_to)
        if not d_from or not d_to:
            return HttpResponse('Invalid date format. Use YYYY-MM-DD', status=400)

        forms_qs = Form.objects.filter(department=dept_id).values_list('id', flat=True)
        if not forms_qs.exists():
            return HttpResponse('No forms for department', status=404)

        apps = Application.objects.filter(
            form_id__in=forms_qs,
            date__gte=d_from,
            date__lt=d_to
        ).select_related('user', 'form')

        wb = Workbook()

        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        bold = Font(bold=True)
        left_align = Alignment(horizontal="left", vertical="center")

        # Лист 1: Статистика по сотрудникам
        ws1 = wb.active
        ws1.title = 'Статистика по сотрудникам'

        ws1.append([dept_name])
        ws1.append([f'Статистика по сотрудникам с {d_from.strftime("%d.%m.%Y")} по {d_to.strftime("%d.%m.%Y")}'])
        ws1.append([])

        ws1["A1"].font = bold
        ws1["A2"].font = bold

        headers1 = ['Исполнитель', 'Принято', 'Выполнено', 'Отклонено', 'Основная форма']
        ws1.append(headers1)
        header_row_idx = ws1.max_row
        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=header_row_idx, column=col_idx)
            cell.font = bold
            cell.alignment = left_align
            cell.border = border

        user_stats = {}
        for a in apps:
            user = a.user.fullname if a.user and a.user.fullname else (a.user.username if a.user else 'Неизвестный')
            usr = user.strip() or 'Неизвестный'
            st = a.status
            form_label = a.form.label if a.form else ''
            entry = user_stats.setdefault(usr, {'Принято': 0, 'Выполнено': 0, 'Отклонено': 0, 'forms': {}})
            if st == ApplicationStatus.IN_PROGRESS:
                entry['Принято'] += 1
            elif st == ApplicationStatus.COMPLETED:
                entry['Выполнено'] += 1
            elif st == ApplicationStatus.REJECTED:
                entry['Отклонено'] += 1
            entry['forms'][form_label] = entry['forms'].get(form_label, 0) + 1

        for user, data in sorted(user_stats.items(), key=lambda x: x[0]):
            main_form = max(data['forms'].items(), key=lambda kv: kv[1])[0] if data['forms'] else ''
            ws1.append([user, data['Принято'], data['Выполнено'], data['Отклонено'], main_form])

        min_row = header_row_idx
        max_row = ws1.max_row
        min_col = 1
        max_col = len(headers1)
        for r in ws1.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in r:
                cell.border = border
                if cell.alignment is None:
                    cell.alignment = left_align

        for col in range(min_col, max_col + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws1[column]:
                if cell.value is None:
                    continue
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            ws1.column_dimensions[column].width = max_length + 2

        # Лист 2: Статистика по формам
        ws2 = wb.create_sheet(title='Статистика по формам')

        ws2.append([dept_name])
        ws2.append([f'Статистика по формам с {d_from.strftime("%d.%m.%Y")} по {d_to.strftime("%d.%m.%Y")}'])
        ws2.append([])

        ws2["A1"].font = bold
        ws2["A2"].font = bold

        headers2 = ['Форма', 'Пришло', 'В работе', 'Выполнено', 'Отменено', 'Отклонено', 'Основной исполнитель']
        ws2.append(headers2)
        header2_row_idx = ws2.max_row
        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=header2_row_idx, column=col_idx)
            cell.font = bold
            cell.alignment = left_align
            cell.border = border

        form_stats = {}
        for a in apps:
            form_label = a.form.label if a.form else 'Без формы'
            st = a.status
            user = a.user.fullname if a.user and a.user.fullname else (a.user.username if a.user else 'Неизвестный')
            entry = form_stats.setdefault(form_label, {'Пришло': 0, 'В работе': 0, 'Выполнено': 0, 'Отменено': 0, 'Отклонено': 0, 'users': {}})
            entry['Пришло'] += 1
            if st == ApplicationStatus.SENDED:
                pass
            if st == ApplicationStatus.IN_PROGRESS:
                entry['В работе'] += 1
            elif st == ApplicationStatus.COMPLETED:
                entry['Выполнено'] += 1
            elif st == ApplicationStatus.CANCELLED:
                entry['Отменено'] += 1
            elif st == ApplicationStatus.REJECTED:
                entry['Отклонено'] += 1
            entry['users'][user] = entry['users'].get(user, 0) + 1

        for form_label, data in sorted(form_stats.items(), key=lambda x: x[0]):
            main_user = max(data['users'].items(), key=lambda kv: kv[1])[0] if data['users'] else ''
            ws2.append([form_label, data['Пришло'], data['В работе'], data['Выполнено'], data['Отменено'], data['Отклонено'], main_user])

        min_row2 = header2_row_idx
        max_row2 = ws2.max_row
        min_col2 = 1
        max_col2 = len(headers2)
        for r in ws2.iter_rows(min_row=min_row2, max_row=max_row2, min_col=min_col2, max_col=max_col2):
            for cell in r:
                cell.border = border
                if cell.alignment is None:
                    cell.alignment = left_align

        for col in range(min_col2, max_col2 + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws2[column]:
                if cell.value is None:
                    continue
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
            ws2.column_dimensions[column].width = max_length + 2

        filename_date_from = d_from.strftime("%d-%m-%Y")
        filename_date_to = d_to.strftime("%d-%m-%Y")
        filename = f'Отдел {dept_name} Статистика с {filename_date_from} по {filename_date_to}.xlsx'
        filename_ascii = iri_to_uri(filename)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename_ascii}"; filename*=UTF-8\'\'{filename_ascii}'
        wb.save(response)
        return response
